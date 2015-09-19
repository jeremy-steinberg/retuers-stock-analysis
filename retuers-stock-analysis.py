from lxml import html
import requests
from openpyxl import Workbook
import datetime
import os

#=============CONSTANTS=================#

REUTERS_BASE_URL = 'http://www.reuters.com/finance/stocks/analyst?symbol='
RATINGS_XPATH = '//td[@class="data dataBold"]/text()'
CONSENSUS_XPATH = '//*[@id="content"]/div[2]/div/div[2]/div[1]/div[2]/div[2]/table/tbody/tr[2]/td[1]/text()'  #taken xpath using chrome inbuilt feature
PREVIOUS_CLOSE_XPATH = '//*[@id="headerQuoteContainer"]/div[3]/div[1]/span[2]/text()'
REUTERS_XLSX_PATH = "reuters.xlsx"

#=================FUNCTIONS=================#

def get_response(stock_name):

        url = REUTERS_BASE_URL + stock_name
        page = requests.get(url)
        xml = html.fromstring(page.text)
        return xml


def import_ratings(stock_name):
        """
        takes a given stock_name and returns the stock data ratings as a table
        """
        return get_response(stock_name).xpath(RATINGS_XPATH)

def import_consensus(stock_name):
        """
        takes a given stock_name and returns the consensus rating
        """
        return get_response(stock_name).xpath(CONSENSUS_XPATH)[0]

def import_prev_close(stock_name):
        """
        Takes a stock name and returns the previous close price
        """
        return get_response(stock_name).xpath(PREVIOUS_CLOSE_XPATH)[0]


def import_excel(stock_name, position):
        """
        Takes a stock name and puts the values in excel and also prints out in the console
        """
        #Takes a stock name, uses the import_ratings function to get the ratings
        #Then puts these ratings on the spreadsheet
        value = import_ratings(stock_name)
        if len(value) < 7:
            print "Stock not found"
            ws[character + str(position)] = float(value[i])
            return
        
        consensus = import_consensus(stock_name)
        ws['A' + str(position)] = stock_name
        
        for i in range(0,7):
                character = chr(ord('B') + i)
                ws[character + str(position)] = float(value[i])
                
        ws['I' + str(position)] = consensus
        ws['J' + str(position)].hyperlink = REUTERS_BASE_URL + stock_name
        
        print "Stock name: \t\t%s" % (stock_name)
        print "Previous close: \t%s" % (import_prev_close(stock_name))
        
        titles = ["Buy(1):\t\t\t", "Outperform(2):\t\t", "Hold(3):\t\t", "Underperform(4):\t",
                  "Sell(5):\t\t", "No opinion(6):\t\t", "Mean:\t\t\t"]
        
        for i in range(len(titles)):
                print "%s %s" %(titles[i], value[i])  

        print "Consensus: \t\t%s"%(consensus)

def spreadsheet_finalise():
        """
        Add titles and save the xlsx
        """
        titles = ["Stocks", "(1)Buy", "(2)Outperform", "(3)Hold", 
                  "(4)Underperform", "(5)Sell", "(6)No opinion", 
                  "Mean", "Consensus", "Link", "File produced: ",
                  "", datetime.datetime.now()]
        
        for i in range(len(titles)):
                character = chr(ord('A') + i)
                ws[character + str(1)] = (titles[i])
                
        ws.merge_cells('M1:N1')
        wb.save(REUTERS_XLSX_PATH)

        print "Spreadsheet saved as reuters.xlsx"




#=============TEST=================#
#test check import_ratings working
#print import_ratings('CNU.NZ')

#test check import_excel working
#import_excel('CNU', 2)

#test check import_consensus working - not working

#============MAIN PROGRAM==========#

# remove current file
if os.path.isfile(REUTERS_XLSX_PATH):
    os.remove(REUTERS_XLSX_PATH)

#Create a spreadsheet and give the sheet the title reuters ratings
wb = Workbook()
ws = wb.active
ws.title = "Reuters ratings"

#introduction and explanation
print "This program imports analyst recommendations from reuters into excel\n"
print "Manual mode allows you to enter each stock individually"
print "Please add the suffixes as appropriate (e.g. .NZ or .AX etc)\n"
print "Import mode takes the stock tickers from the text file stocks.txt"
print "When editing stocks.txt, separate each stock ticker by a comma (no spaces)."
print "e.g. ARG.NZ,CNU.NZ,KPG.NZ,BHP.AX\n"

# Main program. two modes auto (import) and manual.
# Manual: Asks for a stock name, uses the import_excel function on the stock name, then
# asks if you want to add another stock or quit.
# Auto: imports stock names from the text file and then runs the import_excel function on them

to_continue = True
while to_continue:
        choice = raw_input("Choose manual mode (m), import from text file (i) or quit and save to excel (q) ")
        if choice == "m":
                position = 2
                print ""
                while to_continue:
                        Stock_name = raw_input("Choose a stock ticker or quit and save spreadsheet (q) ")
                        print ""
                        if Stock_name == "q":
                                spreadsheet_finalise()
                                to_continue = False
                        else:
                                import_excel(Stock_name, position)
                                position = position + 1
                                print ""
        elif choice == "i":
                print ""
                stocks_file = open("stocks.txt", 'r')
                stocks_list = stocks_file.read().split(',')
                stocks_file.close()
                position = 2
                stock_number = 0
                while stock_number < len(stocks_list): #counts the number of stocks in the array using len
                        import_excel(stocks_list[stock_number], position)
                        position = position + 1
                        stock_number = stock_number + 1
                        print ""
                
                spreadsheet_finalise()
                to_continue = False

        elif choice == "q":
                to_continue = False

        else:
                print "\nNo such option.\n"






##======================================================="

##TODO


##make manual mode and auto mode into their own functions
##work out how to get current price
